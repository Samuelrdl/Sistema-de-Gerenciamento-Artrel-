from flask import Blueprint, jsonify, request, make_response
from src.models.user import (AtribuicaoFerramentaEPI, ServicoExterno, Eletricista, 
                            FerramentaEPI, User, Veiculo, db)
from src.routes.auth import require_auth
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment
import io
from datetime import datetime

export_bp = Blueprint('export', __name__)

# Rotas de busca
@export_bp.route('/search/atribuicoes', methods=['GET'])
@require_auth
def search_atribuicoes():
    eletricista_nome = request.args.get('eletricista_nome', '')
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    item_nome = request.args.get('item_nome', '')
    
    query = db.session.query(AtribuicaoFerramentaEPI).join(Eletricista).join(FerramentaEPI)
    
    if eletricista_nome:
        query = query.filter(Eletricista.nome.ilike(f'%{eletricista_nome}%'))
    
    if data_inicio:
        try:
            data_inicio_dt = datetime.fromisoformat(data_inicio.replace('Z', '+00:00'))
            query = query.filter(AtribuicaoFerramentaEPI.data_retirada >= data_inicio_dt)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim_dt = datetime.fromisoformat(data_fim.replace('Z', '+00:00'))
            query = query.filter(AtribuicaoFerramentaEPI.data_retirada <= data_fim_dt)
        except ValueError:
            pass
    
    if item_nome:
        query = query.filter(FerramentaEPI.nome.ilike(f'%{item_nome}%'))
    
    atribuicoes = query.all()
    return jsonify([atribuicao.to_dict() for atribuicao in atribuicoes])

@export_bp.route('/search/servicos-externos', methods=['GET'])
@require_auth
def search_servicos_externos():
    colaborador_nome = request.args.get('colaborador_nome', '')
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    destino = request.args.get('destino', '')
    empresa = request.args.get('empresa', '')
    
    query = db.session.query(ServicoExterno).join(User)
    
    if colaborador_nome:
        query = query.filter(User.username.ilike(f'%{colaborador_nome}%'))
    
    if data_inicio:
        try:
            data_inicio_dt = datetime.fromisoformat(data_inicio.replace('Z', '+00:00'))
            query = query.filter(ServicoExterno.data_hora_saida >= data_inicio_dt)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim_dt = datetime.fromisoformat(data_fim.replace('Z', '+00:00'))
            query = query.filter(ServicoExterno.data_hora_saida <= data_fim_dt)
        except ValueError:
            pass
    
    if destino:
        query = query.filter(ServicoExterno.destino.ilike(f'%{destino}%'))
    
    if empresa:
        query = query.filter(ServicoExterno.empresa_atendida.ilike(f'%{empresa}%'))
    
    servicos = query.all()
    result = []
    for servico in servicos:
        servico_dict = servico.to_dict()
        servico_dict['materiais'] = [material.to_dict() for material in servico.materiais]
        if servico.checklist_cinto:
            servico_dict['checklist_cinto'] = servico.checklist_cinto.to_dict()
        if servico.checklist_escada:
            servico_dict['checklist_escada'] = servico.checklist_escada.to_dict()
        result.append(servico_dict)
    
    return jsonify(result)

# Rotas de exportação PDF
@export_bp.route('/export/atribuicoes/pdf', methods=['GET'])
@require_auth
def export_atribuicoes_pdf():
    atribuicoes = AtribuicaoFerramentaEPI.query.all()
    
    # Criar PDF em memória
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centralizado
    )
    
    # Conteúdo
    story = []
    
    # Título
    title = Paragraph("Relatório de Atribuições de Ferramentas e EPIs", title_style)
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Dados da tabela
    data = [['Eletricista', 'Item', 'Tipo', 'Data Retirada', 'Data Devolução', 'Observação']]
    
    for atribuicao in atribuicoes:
        data.append([
            atribuicao.eletricista.nome if atribuicao.eletricista else '',
            atribuicao.ferramenta_epi.nome if atribuicao.ferramenta_epi else '',
            atribuicao.ferramenta_epi.tipo if atribuicao.ferramenta_epi else '',
            atribuicao.data_retirada.strftime('%d/%m/%Y %H:%M') if atribuicao.data_retirada else '',
            atribuicao.data_devolucao.strftime('%d/%m/%Y %H:%M') if atribuicao.data_devolucao else 'Não devolvido',
            atribuicao.observacao or ''
        ])
    
    # Criar tabela
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Construir PDF
    doc.build(story)
    
    # Preparar resposta
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=atribuicoes.pdf'
    
    return response

@export_bp.route('/export/servicos-externos/pdf', methods=['GET'])
@require_auth
def export_servicos_externos_pdf():
    servicos = ServicoExterno.query.all()
    
    # Criar PDF em memória
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centralizado
    )
    
    # Conteúdo
    story = []
    
    # Título
    title = Paragraph("Relatório de Serviços Externos", title_style)
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Dados da tabela
    data = [['Colaborador', 'Veículo', 'Destino', 'Empresa', 'Data/Hora Saída']]
    
    for servico in servicos:
        data.append([
            servico.colaborador.username if servico.colaborador else '',
            servico.veiculo.identificacao if servico.veiculo else '',
            servico.destino,
            servico.empresa_atendida,
            servico.data_hora_saida.strftime('%d/%m/%Y %H:%M') if servico.data_hora_saida else ''
        ])
    
    # Criar tabela
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Construir PDF
    doc.build(story)
    
    # Preparar resposta
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=servicos_externos.pdf'
    
    return response

# Rotas de exportação Excel
@export_bp.route('/export/atribuicoes/excel', methods=['GET'])
@require_auth
def export_atribuicoes_excel():
    atribuicoes = AtribuicaoFerramentaEPI.query.all()
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atribuições"
    
    # Cabeçalhos
    headers = ['Eletricista', 'Item', 'Tipo', 'Data Retirada', 'Data Devolução', 'Observação']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Dados
    for row, atribuicao in enumerate(atribuicoes, 2):
        ws.cell(row=row, column=1, value=atribuicao.eletricista.nome if atribuicao.eletricista else '')
        ws.cell(row=row, column=2, value=atribuicao.ferramenta_epi.nome if atribuicao.ferramenta_epi else '')
        ws.cell(row=row, column=3, value=atribuicao.ferramenta_epi.tipo if atribuicao.ferramenta_epi else '')
        ws.cell(row=row, column=4, value=atribuicao.data_retirada.strftime('%d/%m/%Y %H:%M') if atribuicao.data_retirada else '')
        ws.cell(row=row, column=5, value=atribuicao.data_devolucao.strftime('%d/%m/%Y %H:%M') if atribuicao.data_devolucao else 'Não devolvido')
        ws.cell(row=row, column=6, value=atribuicao.observacao or '')
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Preparar resposta
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=atribuicoes.xlsx'
    
    return response

@export_bp.route('/export/servicos-externos/excel', methods=['GET'])
@require_auth
def export_servicos_externos_excel():
    servicos = ServicoExterno.query.all()
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Serviços Externos"
    
    # Cabeçalhos
    headers = ['Colaborador', 'Veículo', 'Destino', 'Empresa', 'Data/Hora Saída']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Dados
    for row, servico in enumerate(servicos, 2):
        ws.cell(row=row, column=1, value=servico.colaborador.username if servico.colaborador else '')
        ws.cell(row=row, column=2, value=servico.veiculo.identificacao if servico.veiculo else '')
        ws.cell(row=row, column=3, value=servico.destino)
        ws.cell(row=row, column=4, value=servico.empresa_atendida)
        ws.cell(row=row, column=5, value=servico.data_hora_saida.strftime('%d/%m/%Y %H:%M') if servico.data_hora_saida else '')
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Preparar resposta
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=servicos_externos.xlsx'
    
    return response

